
import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("📊 Backlog Weekly Check Tool")

# Upload files
prev_file = st.file_uploader("Upload Previous Week's Backlog", type=["xlsx"])
curr_file = st.file_uploader("Upload Current Week's Backlog", type=["xlsx"])
eng_mgr_file = st.file_uploader("Upload Engagement Manager File", type=["xlsx"])
teco_file = st.file_uploader("Upload TECO Status File", type=["xlsx"])

if st.button("Run Comparison") and all([prev_file, curr_file, eng_mgr_file, teco_file]):
    # Load Excel files
    file1 = pd.read_excel(prev_file, engine='openpyxl')
    file2 = pd.read_excel(curr_file, engine='openpyxl')
    eng_mgr = pd.read_excel(eng_mgr_file, engine='openpyxl')
    teco = pd.read_excel(teco_file, engine='openpyxl')

    # Merge and compare
    merged = pd.merge(file1, file2, on=['Sales Order', 'CLI', 'WBS Element'], how='outer', indicator=True)
    new_rows = merged[merged['_merge'] == 'right_only']
    missing_rows = merged[merged['_merge'] == 'left_only']
    comparison = merged[merged['_merge'] == 'both'].copy()

    for column in file1.columns:
        if column not in ['Sales Order', 'CLI', 'WBS Element']:
            comparison[column + '_diff'] = comparison[column + '_x'] != comparison[column + '_y']

    if 'Remaining Backlog_x' in comparison.columns and 'Remaining Backlog_y' in comparison.columns:
        comparison['Remaining Backlog Delta'] = comparison['Remaining Backlog_x'] - comparison['Remaining Backlog_y']

    # Merge with engineering manager
    comparison = pd.merge(
        comparison,
        eng_mgr[['Sales Document', 'Eng Mgr - First name', 'Eng Mgr - Last name']],
        left_on='Sales Order',
        right_on='Sales Document',
        how='left'
    )

    # Drop unnecessary columns
    columns_to_hide = [
        '_merge', 'Sales Organization_diff', 'CLI Start Date_diff', 'CLI End Date_diff',
        'Measurement customer Name 1_diff', 'Item Status_diff', 'Item Net Value LC_diff',
        'Total invoiced_diff', 'Invoiced Currency_diff', 'Remaining Backlog_diff',
        'Contract Currency_diff', 'Sales Document'
    ]
    comparison.drop(columns=[col for col in columns_to_hide if col in comparison.columns], inplace=True)

    # Merge engineering manager info into new and missing rows
    new_rows = pd.merge(
        new_rows,
        eng_mgr[['Sales Document', 'Eng Mgr - First name', 'Eng Mgr - Last name']],
        left_on='Sales Order',
        right_on='Sales Document',
        how='left'
    )
    missing_rows = pd.merge(
        missing_rows,
        eng_mgr[['Sales Document', 'Eng Mgr - First name', 'Eng Mgr - Last name']],
        left_on='Sales Order',
        right_on='Sales Document',
        how='left'
    )
    new_rows.drop(columns=['Sales Document'], errors='ignore', inplace=True)
    missing_rows.drop(columns=['Sales Document'], errors='ignore', inplace=True)

    # Add TECO status to Solved Items
    missing_rows = pd.merge(
        missing_rows,
        teco[['Sales Order', 'WBS Element', 'Item Status']],
        on=['Sales Order', 'WBS Element'],
        how='left'
    )

    # Save to Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        comparison.to_excel(writer, sheet_name='Comparison', index=False)
        new_rows.to_excel(writer, sheet_name='New Items', index=False)
        missing_rows.to_excel(writer, sheet_name='Solved Items', index=False)

    # Highlight deltas
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Comparison']
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Remaining Backlog Delta':
            for cell in col[1:]:
                if cell.value != 0:
                    cell.fill = yellow_fill
            break

    # Save final workbook to BytesIO
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.success("✅ Comparison complete!")
    st.download_button(
        label="📥 Download Result Excel File",
        data=final_output,
        file_name=f"Backlog_analysis_{datetime.now().strftime('%d%m%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
